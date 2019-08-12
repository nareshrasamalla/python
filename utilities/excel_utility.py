import openpyxl

class PCRExcel:

    def __init__(self,workbookLocation,tdh):
        self.workBookLoaction = workbookLocation
        self.tdh = tdh

    def updateCellValue(self,sheetname):
        wb = openpyxl.load_workbook(filename = self.workBookLoaction)
        ws = wb[sheetname]
        #sheet = wb.worksheets(0)
        for j in range(1,int(self.tdh['PCR Excel Sheet - 1-No of Attributes to change'])+1):
            for i in range(1,ws.max_column+1):
                #print(ws.cell(row=1, column=i).value)
                if ws.cell(row=1, column=i).value == self.tdh['PCR Excel Sheet - 1-Column Name - '+str(j)]:
                    ws.cell(row=2, column=i).value = self.tdh['PCR Excel Sheet - 1-Column Value - '+str(j)]

                    break
                else:pass
        wb.save(self.workBookLoaction)
        #wb.save('D:\Python\pythonSelenium\pythonSelenium\\resources\\uiTestData\PCR\\test2.xls')

    def getCertNumber(workbookLocation,sheetName):
        location = (workbookLocation)
        wb = xlrd.open_workbook(location)
        sheet = wb.sheet_by_name(sheetName)
        for i in range(sheet.ncols):
            (sheet.cell(0, i))

